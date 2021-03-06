from bs4 import BeautifulSoup
from selenium import webdriver
from time import sleep
import urllib.request
from openpyxl import load_workbook
import os
import sys
import urllib.request
import json
import re

def search(title):
    client_id = "LOaCf2SwISd49LEn0XSh"
    client_secret = "CdN7ijkQHF"
    encText = urllib.parse.quote(title)
    url = "https://openapi.naver.com/v1/search/local?query=" + encText
    request = urllib.request.Request(url)
    request.add_header("X-Naver-Client-Id", client_id)
    request.add_header("X-Naver-Client-Secret", client_secret)
    response = urllib.request.urlopen(request)
    rescode = response.getcode()
    if(rescode==200):
        response_body = response.read()
        data = response_body.decode('utf-8')
        return data
    else:
        print("Error Code:" + rescode)

def sort(title):
    res = search(title)
    if(res == None):
        print("검색 실패!")
        exit()

    json_response = json.loads(res)
    if(json_response == None):
        print("json.loads 실패!")
        exit()

    # 태그 제거 + value 값 출력
    for item in json_response['items']:
        item["title"] = re.sub(
            "&(.*?);", "", item["title"].replace("<b>", "").replace("</b>", "")
        )
        item["category"] = re.sub(
            "&(.*?);", "", item["category"].replace("<b>", "").replace("</b>", ""),
        )
        if item['category'].startswith('카페,디저트'):
            print("title: %s\t category: %s" %(item['title'], item['category']))

def load_xl():
    wb = load_workbook('거래내역조회20210427.xlsx')
    ws = wb.active

    col = ws['D']
    for cell in col:
        # print(cell.value)
        title = str(cell.value)
        # print(title)
        sort(title)

# def main():
load_xl()
#
# if __name__ == '__main__':
#     main()

